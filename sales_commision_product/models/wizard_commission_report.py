from odoo import api, fields, models
from odoo.exceptions import UserError
from datetime import datetime
import logging
import base64
from io import BytesIO

_logger = logging.getLogger(__name__)


class WizardCommissionReport(models.TransientModel):
    _name = "wizard.commission.report"
    _description = "Commission Financial Report Wizard"

    salesperson_ids = fields.Many2many(
        'res.users',
        string='Salespersons',
        domain=[('share', '=', False)],
        help="Leave empty to include all salespeople"
    )
    date_from = fields.Date(
        string='Date From',
        required=True,
        default=lambda self: fields.Date.today().replace(day=1)
    )
    date_to = fields.Date(
        string='Date To',
        required=True,
        default=fields.Date.today
    )
    status_filter = fields.Selection([
        ('paid', 'Paid Only'),
        ('posted', 'Posted Only (Not Paid)'),
        ('all', 'All Posted Invoices')
    ], string='Invoice Status', required=True, default='posted',
        help="Paid: Only fully paid invoices (actual earnings)\n"
             "Posted Only: Posted but not paid (forecast earnings)\n"
             "All: Both paid and unpaid posted invoices")

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        """Validate date range."""
        for wizard in self:
            if wizard.date_from > wizard.date_to:
                raise UserError("Date From cannot be later than Date To.")

    def _get_commission_data(self):
        """
        Query commission data from sales.commission.line model.
        Returns structured data grouped by salesperson.
        
        This method now queries from the sales.commission.line model (the same
        data source as the Commission Report) instead of rebuilding from account.move,
        ensuring data consistency and better performance.
        """
        self.ensure_one()
        
        # Build domain for commission lines
        domain = [
            ('invoice_date', '>=', self.date_from),
            ('invoice_date', '<=', self.date_to),
        ]
        
        # Status filter - check the related invoice's payment state
        if self.status_filter == 'paid':
            domain.append(('invoice_id.payment_state', 'in', ['paid', 'in_payment']))
        elif self.status_filter == 'posted':
            domain.append(('invoice_id.payment_state', 'not in', ['paid', 'in_payment']))
        # 'all' doesn't need additional payment_state filter
        
        # Ensure invoice is still posted (in case it was cancelled after sync)
        domain.append(('invoice_id.state', '=', 'posted'))
        
        # Salesperson filter
        if self.salesperson_ids:
            domain.append(('salesperson_id', 'in', self.salesperson_ids.ids))
        
        # Get commission lines ordered by salesperson and date
        commission_lines = self.env['sales.commission.line'].search(
            domain, 
            order='salesperson_id, invoice_date'
        )
        
        # Structure: {salesperson_id: {data}}
        data_by_salesperson = {}
        
        for line in commission_lines:
            salesperson = line.salesperson_id
            if not salesperson:
                continue
            
            if salesperson.id not in data_by_salesperson:
                data_by_salesperson[salesperson.id] = {
                    'salesperson': salesperson,
                    'total_sales': 0.0,
                    'total_returns': 0.0,
                    'total_commission': 0.0,
                    'lines': []
                }
            
            # Handle refunds (already negative in commission line)
            if line.move_type == 'out_refund':
                data_by_salesperson[salesperson.id]['total_returns'] += abs(line.line_subtotal)
            else:
                data_by_salesperson[salesperson.id]['total_sales'] += line.line_subtotal
            
            data_by_salesperson[salesperson.id]['total_commission'] += line.commission_amount
            
            # Add line detail
            data_by_salesperson[salesperson.id]['lines'].append({
                'invoice_date': line.invoice_date,
                'invoice_number': line.invoice_id.name,
                'invoice_id': line.invoice_id.id,
                'product_name': line.product_id.display_name,
                'quantity': line.quantity or 0.0,
                'line_subtotal': line.line_subtotal,
                'commission_rate': line.commission_rate or 0.0,
                'commission_amount': line.commission_amount,
                'move_type': 'Invoice' if line.move_type == 'out_invoice' else 'Refund',
            })
        
        # Return as list sorted by salesperson name (for QWeb compatibility)
        # QWeb can't use lambda functions, so we pre-sort here
        sorted_data = [
            {
                'salesperson_id': sp_id,
                **data
            }
            for sp_id, data in sorted(
                data_by_salesperson.items(),
                key=lambda x: x[1]['salesperson'].name
            )
        ]
        
        return sorted_data

    def action_print_excel(self):
        """Generate Excel report and return as download."""
        self.ensure_one()
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError("The 'openpyxl' Python library is required for Excel export. "
                          "Please install it or contact your administrator.")
        
        # Get data
        data = self._get_commission_data()
        
        if not data:
            raise UserError("No commission data found for the selected filters.")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Summary sheet
        ws_summary = wb.create_sheet('Summary')
        
        # Header styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Summary headers
        headers_summary = ['Salesperson', 'Total Sales', 'Total Returns', 'Net Sales', 'Total Commission']
        for col_num, header in enumerate(headers_summary, 1):
            cell = ws_summary.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Summary data
        row_num = 2
        grand_sales = 0.0
        grand_returns = 0.0
        grand_commission = 0.0
        
        # data is now a list, not a dict
        for sp_data in data:
            total_sales = sp_data['total_sales']
            total_returns = sp_data['total_returns']
            net_sales = total_sales - total_returns
            total_commission = sp_data['total_commission']
            
            ws_summary.cell(row=row_num, column=1).value = sp_data['salesperson'].name
            ws_summary.cell(row=row_num, column=2).value = total_sales
            ws_summary.cell(row=row_num, column=2).number_format = '#,##0.00'
            ws_summary.cell(row=row_num, column=3).value = total_returns
            ws_summary.cell(row=row_num, column=3).number_format = '#,##0.00'
            ws_summary.cell(row=row_num, column=4).value = net_sales
            ws_summary.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws_summary.cell(row=row_num, column=5).value = total_commission
            ws_summary.cell(row=row_num, column=5).number_format = '#,##0.00'
            
            grand_sales += total_sales
            grand_returns += total_returns
            grand_commission += total_commission
            
            row_num += 1
        
        # Grand totals
        ws_summary.cell(row=row_num, column=1).value = 'GRAND TOTAL'
        ws_summary.cell(row=row_num, column=1).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=2).value = grand_sales
        ws_summary.cell(row=row_num, column=2).number_format = '#,##0.00'
        ws_summary.cell(row=row_num, column=2).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=3).value = grand_returns
        ws_summary.cell(row=row_num, column=3).number_format = '#,##0.00'
        ws_summary.cell(row=row_num, column=3).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=4).value = grand_sales - grand_returns
        ws_summary.cell(row=row_num, column=4).number_format = '#,##0.00'
        ws_summary.cell(row=row_num, column=4).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=5).value = grand_commission
        ws_summary.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws_summary.cell(row=row_num, column=5).font = Font(bold=True)
        
        # Auto-width columns
        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_summary.column_dimensions[column].width = max_length + 2
        
        # Create Detailed Lines sheet
        ws_detail = wb.create_sheet('Detailed Lines')
        
        # Detail headers
        headers_detail = ['Date', 'Salesperson', 'Invoice', 'Product', 'Quantity', 
                         'Subtotal', 'Commission Rate', 'Commission', 'Type']
        for col_num, header in enumerate(headers_detail, 1):
            cell = ws_detail.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Detail data
        row_num = 2
        # data is now a list, not a dict
        for sp_data in data:
            for line in sp_data['lines']:
                ws_detail.cell(row=row_num, column=1).value = line['invoice_date']
                ws_detail.cell(row=row_num, column=1).number_format = 'YYYY-MM-DD'
                ws_detail.cell(row=row_num, column=2).value = sp_data['salesperson'].name
                ws_detail.cell(row=row_num, column=3).value = line['invoice_number']
                ws_detail.cell(row=row_num, column=4).value = line['product_name']
                ws_detail.cell(row=row_num, column=5).value = line['quantity']
                ws_detail.cell(row=row_num, column=5).number_format = '#,##0.00'
                ws_detail.cell(row=row_num, column=6).value = line['line_subtotal']
                ws_detail.cell(row=row_num, column=6).number_format = '#,##0.00'
                ws_detail.cell(row=row_num, column=7).value = line['commission_rate']
                ws_detail.cell(row=row_num, column=7).number_format = '0.00"%"'
                ws_detail.cell(row=row_num, column=8).value = line['commission_amount']
                ws_detail.cell(row=row_num, column=8).number_format = '#,##0.00'
                ws_detail.cell(row=row_num, column=9).value = line['move_type']
                
                row_num += 1
        
        # Auto-width columns for detail sheet
        for col in ws_detail.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_detail.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create attachment
        filename = f"Commission_Report_{self.date_from}_{self.date_to}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_print_pdf(self):
        """Generate PDF report."""
        self.ensure_one()
        
        # Validate that we have data before generating report
        data = self._get_commission_data()
        
        if not data:
            raise UserError("No commission data found for the selected filters.")
        
        # Generate report - the QWeb template will call _get_commission_data() during rendering
        return self.env.ref('sales_commision_product.action_report_commission_financial').report_action(self)
